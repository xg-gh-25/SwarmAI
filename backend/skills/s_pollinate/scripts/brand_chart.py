#!/usr/bin/env python3
"""
brand_chart.py — Apply Pollinate direction tokens to openpyxl charts.

Reads a direction YAML file and applies its color tokens to all charts in an XLSX file.
Ensures chart series, titles, axes, and legends use brand-consistent styling.

Usage:
    python brand_chart.py <xlsx_path> --direction <direction_yaml> [--json]

No external dependencies beyond openpyxl + pyyaml.
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import yaml
except ImportError:
    print("Error: pyyaml required. pip install pyyaml", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.fill import PatternFillScheme, ColorChoice
except ImportError:
    print("Error: openpyxl required. pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def load_direction_tokens(direction_path: str) -> dict:
    """Load color tokens from direction YAML."""
    path = Path(direction_path)
    if not path.exists():
        raise FileNotFoundError(f"Direction file not found: {path}")

    with open(path) as f:
        data = yaml.safe_load(f)

    return data.get('tokens', {})


def hex_to_openpyxl(hex_color: str) -> str:
    """Convert '#D4A853' or 'D4A853' to openpyxl format 'D4A853' (6-char, no #)."""
    color = hex_color.lstrip('#')
    # Handle rgba() values — extract hex approximation
    if color.startswith('rgba') or color.startswith('rgb'):
        return '808080'  # Fallback gray for rgba values
    # Ensure 6 characters
    if len(color) == 3:
        color = ''.join(c * 2 for c in color)
    return color[:6].upper()


def apply_brand_to_charts(xlsx_path: str, tokens: dict) -> dict:
    """
    Apply brand tokens to all charts in the workbook.

    Returns: {"charts_styled": int, "sheets_processed": int, "errors": []}
    """
    wb = load_workbook(xlsx_path)
    charts_styled = 0
    sheets_processed = 0
    errors = []

    # Extract key colors
    accent = hex_to_openpyxl(tokens.get('accent', '#D4A853'))
    bg_deep = hex_to_openpyxl(tokens.get('bg-deep', '#0A0A0B'))
    text_primary = hex_to_openpyxl(tokens.get('text-primary', '#FFFFFF'))
    text_secondary = hex_to_openpyxl(tokens.get('text-secondary', '#B3B3B3'))

    # Series color palette (accent + neutrals)
    series_colors = [
        accent,          # Primary series
        '4A4A4A',        # Secondary (neutral dark)
        '7A7A7A',        # Tertiary (neutral medium)
        text_secondary.replace('rgba', '').strip('()'),  # Quaternary
    ]

    for ws in wb.worksheets:
        sheets_processed += 1
        for chart in ws._charts:
            try:
                # Apply series colors
                for i, series in enumerate(chart.series):
                    color = series_colors[i % len(series_colors)]
                    try:
                        series.graphicalProperties.solidFill = color
                    except (AttributeError, TypeError):
                        pass  # Some chart types don't support direct fill

                # Style title
                if chart.title:
                    try:
                        chart.title.txPr = None  # Reset to allow styling
                    except AttributeError:
                        pass

                # Legend position
                if chart.legend:
                    chart.legend.position = 'b'  # Bottom

                # Minimal gridlines
                try:
                    chart.y_axis.majorGridlines = None
                    chart.y_axis.minorGridlines = None
                except AttributeError:
                    pass

                charts_styled += 1
            except Exception as e:
                errors.append(f"Chart in '{ws.title}': {str(e)}")

    # Save
    wb.save(xlsx_path)

    return {
        "charts_styled": charts_styled,
        "sheets_processed": sheets_processed,
        "accent_applied": accent,
        "errors": errors,
        "success": len(errors) == 0,
    }


def main():
    parser = argparse.ArgumentParser(description="Apply brand tokens to XLSX charts")
    parser.add_argument("xlsx_path", help="Path to XLSX file")
    parser.add_argument("--direction", required=True, help="Path to direction YAML")
    parser.add_argument("--json", action="store_true", help="Output as JSON")
    args = parser.parse_args()

    if not Path(args.xlsx_path).exists():
        print(f"Error: {args.xlsx_path} not found", file=sys.stderr)
        sys.exit(1)

    tokens = load_direction_tokens(args.direction)
    result = apply_brand_to_charts(args.xlsx_path, tokens)

    if args.json:
        print(json.dumps(result, indent=2))
    else:
        status = "✅" if result["success"] else "⚠️"
        print(f"{status} Styled {result['charts_styled']} charts across {result['sheets_processed']} sheets")
        print(f"   Accent color: #{result['accent_applied']}")
        if result['errors']:
            for e in result['errors']:
                print(f"   ERROR: {e}")

    sys.exit(0 if result["success"] else 1)


if __name__ == "__main__":
    main()
