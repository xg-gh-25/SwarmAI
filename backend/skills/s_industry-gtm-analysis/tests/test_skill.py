"""Tests for s_industry-gtm-analysis skill — TDD RED phase."""
import pytest
import json
import sys
import os

# Add skill to path
SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, SKILL_DIR)


class TestDataModule:
    """AC1: data.py queries revenue from Athena for any BU."""

    def test_query_revenue_returns_list(self):
        from data import query_revenue
        result = query_revenue(bu_name="AUTO & MFG", tshirt_min="L")
        assert isinstance(result, list)

    def test_query_revenue_has_required_fields(self):
        from data import query_revenue
        result = query_revenue(bu_name="AUTO & MFG", tshirt_min="L")
        if result:
            row = result[0]
            for field in ['sfdc_account_name', 'sfdc_account_18id', 'ttm_revenue', 'genai_ttm', 'bedrock_ttm']:
                assert field in row, f"Missing field: {field}"

    def test_query_revenue_filters_by_bu(self):
        from data import query_revenue
        # Should not raise for any BU name
        result = query_revenue(bu_name="FSI-DNB", tshirt_min="XL")
        assert isinstance(result, list)

    def test_build_sfdc_url(self):
        from data import build_sfdc_url
        url = build_sfdc_url("0015000000uWCH8AAO")
        assert "aws-crm.lightning.force.com" in url
        assert "0015000000uWCH8AAO" in url


class TestExcelBuilder:
    """AC2: Excel builder generates multi-sheet workbook."""

    def test_build_excel_creates_file(self, tmp_path):
        from templates.excel_builder import build_excel
        accounts = [
            {"short": "TestCo", "name": "Test Company", "website": "test.com",
             "size": "XL", "owner": "Owner", "category": "消费电子",
             "ttm": 1000000, "genai": 50000, "bedrock": 10000,
             "products": "Test products", "ai_scenarios": "AI test",
             "openclaw": "No", "maturity": "Medium", "gtm": "Test",
             "sfdc_url": "https://example.com", "industry": "MFG"}
        ]
        outpath = tmp_path / "test.xlsx"
        build_excel(accounts, str(outpath), product="agentcore")
        assert outpath.exists()
        assert outpath.stat().st_size > 0

    def test_build_excel_has_auto_and_mfg_sheets(self, tmp_path):
        from templates.excel_builder import build_excel
        from openpyxl import load_workbook
        accounts = [
            {"short": "AutoCo", "name": "Auto Co", "website": "a.com",
             "size": "XL", "owner": "O", "category": "汽车整车",
             "ttm": 100, "genai": 0, "bedrock": 0,
             "products": "Cars", "ai_scenarios": "AI", "openclaw": "No",
             "maturity": "Low", "gtm": "", "sfdc_url": "", "industry": "AUTO"},
            {"short": "MfgCo", "name": "Mfg Co", "website": "m.com",
             "size": "L", "owner": "O", "category": "工业制造/重工",
             "ttm": 200, "genai": 0, "bedrock": 0,
             "products": "Machines", "ai_scenarios": "AI", "openclaw": "No",
             "maturity": "Low", "gtm": "", "sfdc_url": "", "industry": "MFG"},
        ]
        outpath = tmp_path / "test.xlsx"
        build_excel(accounts, str(outpath), product="agentcore")
        wb = load_workbook(str(outpath))
        sheet_names = wb.sheetnames
        assert any("AUTO" in s for s in sheet_names)
        assert any("MFG" in s for s in sheet_names)


class TestRobSummary:
    """AC3: Rob summary generates 1-page Chinese HTML."""

    def test_generates_html(self):
        from templates.rob_summary import render_rob_summary
        data = {
            "bu_name": "AUTO & MFG",
            "product": "AgentCore",
            "total_accounts": 202,
            "total_ttm": 61684232,
            "total_genai": 3900216,
            "total_bedrock": 1451904,
            "top_accounts": [{"name": "TP-LINK", "ttm": 5184606, "genai": 2413, "bedrock": 1033}],
            "bedrock_accounts": [{"name": "理想汽车", "bedrock": 396828}],
        }
        html = render_rob_summary(data)
        assert "<!DOCTYPE html>" in html
        assert "机密" in html or "AWS" in html
        assert "AgentCore" in html

    def test_html_is_chinese(self):
        from templates.rob_summary import render_rob_summary
        data = {"bu_name": "AUTO & MFG", "product": "AgentCore",
                "total_accounts": 10, "total_ttm": 1000, "total_genai": 100,
                "total_bedrock": 50, "top_accounts": [], "bedrock_accounts": []}
        html = render_rob_summary(data)
        assert "总览" in html or "客户" in html or "分析" in html


class TestGmDetailed:
    """AC4: GM detailed generates multi-tab Chinese HTML."""

    def test_generates_html_with_tabs(self):
        from templates.gm_detailed import render_gm_detailed
        data = {
            "bu_name": "AUTO & MFG", "product": "AgentCore",
            "total_accounts": 202, "total_ttm": 61684232,
            "total_genai": 3900216, "total_bedrock": 1451904,
            "auto_count": 46, "mfg_count": 156,
            "top_accounts": [{"name": "TP-LINK", "ttm": 5184606, "genai": 2413, "bedrock": 1033}],
            "bedrock_accounts": [{"name": "理想汽车", "bedrock": 396828}],
            "categories": {"消费电子": {"count": 40, "ttm": 18000000}},
            "plays": [], "scenarios": [], "competitive": [],
        }
        html = render_gm_detailed(data)
        assert "tab" in html.lower() or "input" in html.lower()
        assert "<!DOCTYPE html>" in html


class TestAnalyzer:
    """AC5: analyzer.py CLI orchestrates the full flow."""

    def test_analyzer_imports(self):
        import analyzer
        assert hasattr(analyzer, 'main') or hasattr(analyzer, 'run_analysis')

    def test_analyzer_accepts_bu_and_product(self):
        import analyzer
        # Should have argument parsing
        assert hasattr(analyzer, 'parse_args') or hasattr(analyzer, 'run_analysis')


class TestSkillMd:
    """AC6: SKILL.md exists with correct triggers."""

    def test_skill_md_exists(self):
        skill_md = os.path.join(SKILL_DIR, "SKILL.md")
        assert os.path.exists(skill_md)

    def test_skill_md_has_trigger(self):
        skill_md = os.path.join(SKILL_DIR, "SKILL.md")
        with open(skill_md) as f:
            content = f.read()
        assert "GTM" in content or "gtm" in content
        assert "TRIGGER" in content or "trigger" in content.lower()
