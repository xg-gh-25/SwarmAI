"""
Working-level Excel builder -- multi-sheet customer analysis workbook.
Sheets: AUTO 客户明细 / MFG 客户明细 / GTM Plays / AgentCore vs 竞品 /
        AI場景 by 産品大類 / Bedrock 渗透分析
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles ────────────────────────────────────────────────────────────────
HF = PatternFill('solid', fgColor='1F4E79')
HFont = Font(bold=True, color='FFFFFF', size=11)
XXL_F = PatternFill('solid', fgColor='FFF2CC')
XL_F = PatternFill('solid', fgColor='E2EFDA')
L_F = PatternFill('solid', fgColor='D6E4F0')
AUTO_BG = PatternFill('solid', fgColor='D6E4F0')   # Blue background for AUTO
MFG_BG = PatternFill('solid', fgColor='E2EFDA')    # Green background for MFG
ZEBRA_F = PatternFill('solid', fgColor='F2F2F2')    # Alternating row fill
AC_COL_F = PatternFill('solid', fgColor='E2EFDA')   # AgentCore column highlight
TITLE_Font = Font(bold=True, color='1F4E79', size=14)
SECTION_Font = Font(bold=True, color='1F4E79', size=12)
GREEN_F = Font(color='006100')                        # ✅ green
RED_F = Font(color='C00000')                          # ❌ red
YELLOW_F = Font(color='BF8F00')                       # 有限 yellow/amber
C = Alignment(horizontal='center', vertical='center', wrap_text=True)
W = Alignment(wrap_text=True, vertical='top')
T = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
           top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

AUTO_CATS = {'汽车整车', '自动驾驶/智驾', '汽车零部件', '出行/两轮'}


# ── Sheet 1/2: Account detail ────────────────────────────────────────────

def _write_sheet(ws, accounts):
    """Write account data to a worksheet."""
    headers = ['#', '客户名称', '全称', '网站', 'Size', 'Owner',
               'TTM Revenue', 'GenAI TTM', 'Bedrock TTM', 'GenAI %',
               '产品大类', '主要产品', 'AI Agent 场景', 'OpenClaw',
               'AI成熟度', 'GTM备注', 'SFDC Link']
    widths = [4, 15, 28, 20, 6, 13, 12, 11, 11, 8,
              13, 26, 30, 24, 9, 26, 14]

    for c_idx, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font = HFont
        cell.fill = HF
        cell.alignment = C
        cell.border = T
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    ws.freeze_panes = 'A2'

    SIZE_ORD = {'XXL': 0, 'XL': 1, 'L': 2}
    accounts.sort(key=lambda x: (SIZE_ORD.get(x.get('size', ''), 9), x.get('short', '')))

    for i, acc in enumerate(accounts, 1):
        r = i + 1
        ttm = float(acc.get('ttm', 0) or 0)
        genai = float(acc.get('genai', 0) or 0)
        bedrock = float(acc.get('bedrock', 0) or 0)
        genai_pct = genai / ttm if ttm > 0 else 0

        ws.cell(row=r, column=1, value=i).alignment = C
        ws.cell(row=r, column=2, value=acc.get('short', ''))
        ws.cell(row=r, column=3, value=acc.get('name', ''))
        ws.cell(row=r, column=4, value=acc.get('website', ''))
        ws.cell(row=r, column=5, value=acc.get('size', '')).alignment = C
        ws.cell(row=r, column=6, value=acc.get('owner', ''))
        ws.cell(row=r, column=7, value=round(ttm)).number_format = '#,##0'
        ws.cell(row=r, column=7).alignment = C
        ws.cell(row=r, column=8, value=round(genai)).number_format = '#,##0'
        ws.cell(row=r, column=8).alignment = C
        ws.cell(row=r, column=9, value=round(bedrock)).number_format = '#,##0'
        ws.cell(row=r, column=9).alignment = C
        ws.cell(row=r, column=10, value=genai_pct).number_format = '0.0%'
        ws.cell(row=r, column=10).alignment = C
        ws.cell(row=r, column=11, value=acc.get('category', ''))
        ws.cell(row=r, column=12, value=acc.get('products', '')).alignment = W
        ws.cell(row=r, column=13, value=acc.get('ai_scenarios', '')).alignment = W
        ws.cell(row=r, column=14, value=acc.get('openclaw', '')).alignment = W
        ws.cell(row=r, column=15, value=acc.get('maturity', '')).alignment = C
        ws.cell(row=r, column=16, value=acc.get('gtm', '')).alignment = W

        sfdc = acc.get('sfdc_url', '')
        if sfdc:
            cell = ws.cell(row=r, column=17)
            cell.value = sfdc
            cell.hyperlink = sfdc
            cell.font = Font(color='0563C1', underline='single', size=9)

        sf = {'XXL': XXL_F, 'XL': XL_F, 'L': L_F}.get(acc.get('size', ''))
        if sf:
            ws.cell(row=r, column=5).fill = sf
        if bedrock > 0:
            ws.cell(row=r, column=9).fill = PatternFill('solid', fgColor='E2EFDA')

        for c_idx in range(1, 18):
            ws.cell(row=r, column=c_idx).border = T

    ws.auto_filter.ref = f'A1:Q{len(accounts) + 1}'
    return len(accounts)


# ── Sheet 3: GTM Plays (Gap 2) ───────────────────────────────────────────

def _write_gtm_plays_sheet(ws, gtm_plays_data: dict, product: str):
    """Write the GTM Plays sheet with 4 sections matching the manual version."""
    product_title = product.title() if product else "Product"

    # -- Title row --
    ws.cell(row=1, column=1, value=f'{product_title} GTM Plays -- AUTO & MFG (v3)')
    ws.cell(row=1, column=1).font = TITLE_Font
    ws.merge_cells('A1:H1')

    # ═══════════════ Section 1: 6 GTM Plays (rows 3-9) ═══════════════
    play_headers = ['GTM Play', 'AgentCore 组件', '客户痛点', '为什么AgentCore',
                    '对外产品场景', '内部企业场景', '优先客户', '类型/难度']
    play_widths = [18, 18, 18, 22, 22, 18, 22, 12]

    for c_idx, (h, w) in enumerate(zip(play_headers, play_widths), 1):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.font = HFont
        cell.fill = HF
        cell.alignment = C
        cell.border = T
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    plays = gtm_plays_data.get("plays", [])
    for i, play in enumerate(plays):
        r = 4 + i
        fields = [play.get("name", ""), play.get("components", ""),
                  play.get("pain", ""), play.get("why", ""),
                  play.get("external", ""), play.get("internal", ""),
                  play.get("customers", ""), play.get("type_diff", "")]
        for c_idx, val in enumerate(fields, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.alignment = W
            cell.border = T

    # ═══════════════ Section 2: 对外产品场景 (rows 11+) ═══════════════
    ext_start = 4 + max(len(plays), 1) + 2  # blank row after plays
    ws.cell(row=ext_start, column=1,
            value='对外产品AI赋能 -- 19场景 (AUTO蓝底 / MFG绿底)')
    ws.cell(row=ext_start, column=1).font = SECTION_Font
    ws.merge_cells(start_row=ext_start, start_column=1,
                   end_row=ext_start, end_column=8)

    ext_headers = ['行业', '价值', '场景', 'AgentCore组件',
                   '客户+Size', '类型', '难度', '关联Play']
    hdr_row = ext_start + 1
    for c_idx, h in enumerate(ext_headers, 1):
        cell = ws.cell(row=hdr_row, column=c_idx, value=h)
        cell.font = HFont
        cell.fill = HF
        cell.alignment = C
        cell.border = T

    ext_scenarios = gtm_plays_data.get("external_scenarios", [])
    for i, s in enumerate(ext_scenarios):
        r = hdr_row + 1 + i
        fields = [s.get("industry", ""), s.get("value", ""), s.get("name", ""),
                  s.get("components", ""), s.get("customers", ""),
                  s.get("type", ""), s.get("difficulty", ""), s.get("play", "")]
        is_auto = s.get("industry") == "AUTO"
        row_fill = AUTO_BG if is_auto else MFG_BG
        for c_idx, val in enumerate(fields, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.alignment = W
            cell.border = T
            if c_idx == 1:
                cell.fill = row_fill

    # ═══════════════ Section 3: 内部企业Agent场景 ═══════════════
    int_start = hdr_row + 1 + max(len(ext_scenarios), 1) + 2
    ws.cell(row=int_start, column=1,
            value='内部企业Agent场景 (AUTO & MFG 通用)')
    ws.cell(row=int_start, column=1).font = SECTION_Font
    ws.merge_cells(start_row=int_start, start_column=1,
                   end_row=int_start, end_column=8)

    int_headers = ['场景', 'Agent做什么', 'vs传统', 'AgentCore组件',
                   '灯塔(AUTO)', '灯塔(MFG)', 'Why Cloud', '关联Play']
    int_hdr_row = int_start + 1
    for c_idx, h in enumerate(int_headers, 1):
        cell = ws.cell(row=int_hdr_row, column=c_idx, value=h)
        cell.font = HFont
        cell.fill = HF
        cell.alignment = C
        cell.border = T

    int_scenarios = gtm_plays_data.get("internal_scenarios", [])
    for i, s in enumerate(int_scenarios):
        r = int_hdr_row + 1 + i
        fields = [s.get("name", ""), s.get("agent_does", ""),
                  s.get("vs_traditional", ""), s.get("components", ""),
                  s.get("auto_lighthouse", ""), s.get("mfg_lighthouse", ""),
                  s.get("why_cloud", ""), s.get("play", "")]
        for c_idx, val in enumerate(fields, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.alignment = W
            cell.border = T

    # ═══════════════ Section 4: Q2 POC 推荐 ═══════════════
    poc_start = int_hdr_row + 1 + max(len(int_scenarios), 1) + 2
    ws.cell(row=poc_start, column=1,
            value='Q2 推荐 3+2 POC')
    ws.cell(row=poc_start, column=1).font = SECTION_Font
    ws.merge_cells(start_row=poc_start, start_column=1,
                   end_row=poc_start, end_column=8)

    poc_headers = ['#', 'POC', '客户', '行业', 'AgentCore组件',
                   '为什么', '类型', '目标']
    poc_hdr_row = poc_start + 1
    for c_idx, h in enumerate(poc_headers, 1):
        cell = ws.cell(row=poc_hdr_row, column=c_idx, value=h)
        cell.font = HFont
        cell.fill = HF
        cell.alignment = C
        cell.border = T

    poc_recs = gtm_plays_data.get("poc_recommendations", [])
    for i, p in enumerate(poc_recs):
        r = poc_hdr_row + 1 + i
        fields = [p.get("num", ""), p.get("poc", ""), p.get("customer", ""),
                  p.get("industry", ""), p.get("components", ""),
                  p.get("why", ""), p.get("type", ""), p.get("target", "")]
        for c_idx, val in enumerate(fields, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.alignment = W
            cell.border = T


# ── Sheet 4: 竞品对比 (Gap 1) ────────────────────────────────────────────

def _competitive_font(value: str) -> Font:
    """Pick font color based on cell value: green for ✅, red for ❌, amber for 有限."""
    if not value:
        return Font(size=10)
    if '✅' in value:
        return GREEN_F
    if '❌' in value:
        return RED_F
    if '有限' in value or '插件' in value or '通义' in value or '豆包' in value or '盘古' in value or '华为账号' in value:
        return YELLOW_F
    return Font(size=10)


def _write_competitive_sheet(ws, competitive_data: dict, product: str):
    """Write the AgentCore vs 竞品 sheet."""
    product_title = product.title() if product else "Product"
    competitors = competitive_data.get("competitors", [])
    rows = competitive_data.get("rows", [])

    if not competitors or not rows:
        ws.cell(row=1, column=1, value="No competitive data available for this product")
        return

    # Title row (merged)
    ws.cell(row=1, column=1, value=f'{product_title} 组件级竞品对比')
    ws.cell(row=1, column=1).font = TITLE_Font
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(competitors) + 1)

    # Header row (row 3, row 2 blank)
    col_headers = [f'{product_title} 组件'] + competitors
    col_widths = [20] + [22] + [10] * (len(competitors) - 1)

    for c_idx, (h, w) in enumerate(zip(col_headers, col_widths), 1):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.font = HFont
        cell.fill = HF
        cell.alignment = C
        cell.border = T
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    # Data rows
    for i, row_data in enumerate(rows):
        r = 4 + i
        # Dimension name in column 1
        ws.cell(row=r, column=1, value=row_data["dimension"]).alignment = W
        ws.cell(row=r, column=1).border = T
        ws.cell(row=r, column=1).font = Font(bold=True, size=10)

        # Competitor values in columns 2+
        for j, val in enumerate(row_data["values"]):
            cell = ws.cell(row=r, column=j + 2, value=val)
            cell.alignment = C
            cell.border = T
            cell.font = _competitive_font(val)

            # AgentCore column (col 2) gets highlight background
            if j == 0:
                cell.fill = AC_COL_F

            # Zebra striping for other columns
            if j > 0 and i % 2 == 0:
                cell.fill = ZEBRA_F

    ws.freeze_panes = 'B4'


# ── Sheet 5: AI場景 by 産品大類 (Gaps 3+4) ────────────────────────────────

def _write_category_sheet(ws, categories: dict, scenario_rows: list = None):
    """Sheet: AI場景 by 産品大類 -- expanded with 13 columns and multi-row per category.

    If scenario_rows is provided (Gap 3+4), writes the expanded 13-column version
    with merged cells for category-level data. Otherwise falls back to the
    original 8-column format.
    """
    if scenario_rows:
        return _write_expanded_category_sheet(ws, scenario_rows)
    return _write_simple_category_sheet(ws, categories)


def _write_simple_category_sheet(ws, categories: dict):
    """Original 8-column category sheet (fallback)."""
    headers = ['行业', '产品大类', '客户数', 'TTM', 'GenAI', 'Bedrock', 'XXL客户', 'XL客户']
    widths = [8, 18, 8, 14, 12, 12, 30, 30]

    for c_idx, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font = HFont
        cell.fill = HF
        cell.alignment = C
        cell.border = T
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    ws.freeze_panes = 'A2'

    sorted_cats = sorted(
        categories.items(),
        key=lambda x: (0 if x[1].get('industry') == 'AUTO' else 1, -x[1].get('ttm', 0))
    )

    for i, (cat_name, cat_data) in enumerate(sorted_cats, 1):
        r = i + 1
        industry = cat_data.get('industry', 'MFG')
        is_auto = industry == 'AUTO'
        row_fill = AUTO_BG if is_auto else MFG_BG

        ws.cell(row=r, column=1, value=industry).alignment = C
        ws.cell(row=r, column=1).fill = row_fill
        ws.cell(row=r, column=2, value=cat_name)
        ws.cell(row=r, column=3, value=cat_data.get('count', 0)).alignment = C
        ws.cell(row=r, column=4, value=round(cat_data.get('ttm', 0))).number_format = '#,##0'
        ws.cell(row=r, column=4).alignment = C
        ws.cell(row=r, column=5, value=round(cat_data.get('genai', 0))).number_format = '#,##0'
        ws.cell(row=r, column=5).alignment = C
        ws.cell(row=r, column=6, value=round(cat_data.get('bedrock', 0))).number_format = '#,##0'
        ws.cell(row=r, column=6).alignment = C
        ws.cell(row=r, column=7, value=', '.join(cat_data.get('xxl', []))).alignment = W
        ws.cell(row=r, column=8, value=', '.join(cat_data.get('xl', []))).alignment = W

        for c_idx in range(1, 9):
            ws.cell(row=r, column=c_idx).border = T

    ws.auto_filter.ref = f'A1:H{len(sorted_cats) + 1}'
    return len(sorted_cats)


def _write_expanded_category_sheet(ws, scenario_rows: list):
    """Expanded 13-column AI場景 sheet with merged cells per category.

    Columns: 行業 | 産品大類 | 客戸数 | TTM Revenue | GenAI TTM | Bedrock TTM |
             XXL客戸 | XL客戸 | 場景類型 | 場景名称 | AgentCore組件 |
             推理依拠 | 類型+理由
    """
    headers = ['行业', '产品大类', '客户数', 'TTM Revenue', 'GenAI TTM', 'Bedrock TTM',
               'XXL客户', 'XL客户', '场景类型', '场景名称', 'AgentCore组件',
               '推理依据(为什么是机会+调研证据)', '类型+理由']
    widths = [5, 12, 5, 12, 11, 11, 18, 28, 7, 18, 18, 45, 20]

    for c_idx, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font = HFont
        cell.fill = HF
        cell.alignment = C
        cell.border = T
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    ws.freeze_panes = 'I2'

    r = 2  # current row
    for row_data in scenario_rows:
        is_first = row_data.get("is_first_in_category", False)
        cat_span = row_data.get("category_row_count", 1)
        industry = row_data.get("industry", "MFG")
        is_auto = industry == "AUTO"
        row_fill = AUTO_BG if is_auto else MFG_BG

        if is_first:
            # Write category-level data (columns 1-8) on the first row
            # Then merge these columns across cat_span rows
            ws.cell(row=r, column=1, value=industry).alignment = C
            ws.cell(row=r, column=1).fill = row_fill
            ws.cell(row=r, column=2, value=row_data.get("category", ""))
            ws.cell(row=r, column=3, value=row_data.get("count", 0)).alignment = C
            ws.cell(row=r, column=4, value=round(row_data.get("ttm", 0))).number_format = '#,##0'
            ws.cell(row=r, column=4).alignment = C
            ws.cell(row=r, column=5, value=round(row_data.get("genai", 0))).number_format = '#,##0'
            ws.cell(row=r, column=5).alignment = C
            ws.cell(row=r, column=6, value=round(row_data.get("bedrock", 0))).number_format = '#,##0'
            ws.cell(row=r, column=6).alignment = C
            ws.cell(row=r, column=7, value=row_data.get("xxl", "")).alignment = W
            ws.cell(row=r, column=8, value=row_data.get("xl", "")).alignment = W

            # Merge columns 1-8 across the category row span
            if cat_span > 1:
                for merge_col in range(1, 9):
                    ws.merge_cells(
                        start_row=r, start_column=merge_col,
                        end_row=r + cat_span - 1, end_column=merge_col
                    )

        # Write scenario-level data (columns 9-13) on every row
        ws.cell(row=r, column=9, value=row_data.get("scene_type", "")).alignment = C
        ws.cell(row=r, column=10, value=row_data.get("scene_name", "")).alignment = W
        ws.cell(row=r, column=11, value=row_data.get("components", "")).alignment = W
        ws.cell(row=r, column=12, value=row_data.get("reasoning", "")).alignment = W
        ws.cell(row=r, column=13, value=row_data.get("verdict", "")).alignment = W

        # Borders for all 13 columns
        for c_idx in range(1, 14):
            ws.cell(row=r, column=c_idx).border = T

        r += 1

    total_rows = len(scenario_rows)
    ws.auto_filter.ref = f'A1:M{total_rows + 1}'
    return total_rows


# ── Sheet 6: Bedrock 渗透分析 ─────────────────────────────────────────────

def _write_bedrock_sheet(ws, accounts: list[dict]):
    """Sheet: Bedrock 渗透分析 -- accounts with Bedrock > 0, sorted desc."""
    headers = ['#', '客户', 'Size', 'TTM', 'Bedrock TTM', 'Bedrock %', '产品大类']
    widths = [4, 18, 6, 14, 14, 10, 18]

    for c_idx, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font = HFont
        cell.fill = HF
        cell.alignment = C
        cell.border = T
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    ws.freeze_panes = 'A2'

    # Filter and sort by bedrock desc
    bedrock_accts = [a for a in accounts if float(a.get('bedrock', 0) or 0) > 0]
    bedrock_accts.sort(key=lambda x: -float(x.get('bedrock', 0) or 0))

    for i, acc in enumerate(bedrock_accts, 1):
        r = i + 1
        ttm = float(acc.get('ttm', 0) or 0)
        bedrock = float(acc.get('bedrock', 0) or 0)
        bedrock_pct = bedrock / ttm if ttm > 0 else 0

        ws.cell(row=r, column=1, value=i).alignment = C
        ws.cell(row=r, column=2, value=acc.get('short', acc.get('name', '')))
        ws.cell(row=r, column=3, value=acc.get('size', '')).alignment = C
        ws.cell(row=r, column=4, value=round(ttm)).number_format = '#,##0'
        ws.cell(row=r, column=4).alignment = C
        ws.cell(row=r, column=5, value=round(bedrock)).number_format = '#,##0'
        ws.cell(row=r, column=5).alignment = C
        ws.cell(row=r, column=5).fill = PatternFill('solid', fgColor='E2EFDA')
        ws.cell(row=r, column=6, value=bedrock_pct).number_format = '0.0%'
        ws.cell(row=r, column=6).alignment = C
        ws.cell(row=r, column=7, value=acc.get('category', ''))

        # Size fill
        sf = {'XXL': XXL_F, 'XL': XL_F, 'L': L_F}.get(acc.get('size', ''))
        if sf:
            ws.cell(row=r, column=3).fill = sf

        for c_idx in range(1, 8):
            ws.cell(row=r, column=c_idx).border = T

    ws.auto_filter.ref = f'A1:G{len(bedrock_accts) + 1}'
    return len(bedrock_accts)


# ── Main entry point ──────────────────────────────────────────────────────

def build_excel(accounts: list[dict], output_path: str, product: str = "",
                categories: dict = None, bedrock_penetration: dict = None,
                competitive: dict = None, gtm_plays_data: dict = None,
                scenario_rows: list = None):
    """
    Build multi-sheet Excel workbook.

    Args:
        accounts: list of account dicts with revenue data
        output_path: where to save the .xlsx file
        product: AWS product name (for sheet titles)
        categories: dict of category distributions (from analyzer)
        bedrock_penetration: dict with penetration stats (from analyzer)
        competitive: dict with competitors and rows for 竞品 sheet
        gtm_plays_data: dict with plays/external_scenarios/internal_scenarios/poc
        scenario_rows: list of expanded scenario row dicts for AI場景 sheet
    """
    wb = Workbook()

    # Split AUTO vs MFG
    auto = [a for a in accounts if a.get('category', '') in AUTO_CATS or a.get('industry') == 'AUTO']
    mfg = [a for a in accounts if a not in auto]

    # Sheet 1: AUTO
    ws_auto = wb.active
    ws_auto.title = 'AUTO 客户明细'
    n_auto = _write_sheet(ws_auto, auto)

    # Sheet 2: MFG
    ws_mfg = wb.create_sheet('MFG 客户明细')
    n_mfg = _write_sheet(ws_mfg, mfg)

    # Sheet 3: GTM Plays (Gap 2)
    if gtm_plays_data and gtm_plays_data.get("plays"):
        ws_plays = wb.create_sheet('GTM Plays')
        _write_gtm_plays_sheet(ws_plays, gtm_plays_data, product)

    # Sheet 4: 竞品対比 (Gap 1)
    if competitive and competitive.get("rows"):
        product_title = product.title() if product else "Product"
        ws_comp = wb.create_sheet(f'{product_title} vs 竞品')
        _write_competitive_sheet(ws_comp, competitive, product)

    # Sheet 5: AI場景 by 産品大類 (Gaps 3+4)
    if categories or scenario_rows:
        ws_cat = wb.create_sheet('AI场景 by 产品大类')
        _write_category_sheet(ws_cat, categories, scenario_rows)

    # Sheet 6: Bedrock 渗透分析
    bedrock_accts = [a for a in accounts if float(a.get('bedrock', 0) or 0) > 0]
    if bedrock_accts:
        ws_br = wb.create_sheet('Bedrock 渗透分析')
        _write_bedrock_sheet(ws_br, accounts)

    wb.save(output_path)
    return {"auto": n_auto, "mfg": n_mfg, "total": n_auto + n_mfg}
